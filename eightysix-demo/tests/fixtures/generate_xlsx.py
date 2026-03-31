"""Generate a multi-tab XLSX that mimics a real messy restaurant export."""

import random
from datetime import date, timedelta, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment

OUTPUT = Path(__file__).parent / "multi_tab_export.xlsx"


def main():
    wb = openpyxl.Workbook()

    # Tab 1: Daily Sales
    ws1 = wb.active
    ws1.title = "Daily Sales"
    ws1.merge_cells("A1:E1")
    ws1["A1"] = "Demo Grill — Q1 2025 Sales Report"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1.append([])  # blank row
    ws1.append(["Business Date", "Gross Sales", "Net Sales", "Order Count", "Delivery $"])

    start = date(2025, 1, 1)
    for i in range(90):
        d = start + timedelta(days=i)
        base = {0: 3200, 1: 2900, 2: 3100, 3: 3400, 4: 4200, 5: 4800, 6: 3600}
        net = base[d.weekday()] + random.randint(-400, 400)
        gross = round(net * 1.04, 2)
        orders = int(net / random.uniform(18, 28))
        delivery = round(net * random.uniform(0.15, 0.35), 2)
        ws1.append([d.strftime("%m/%d/%Y"), f"${gross:,.2f}", f"${net:,.2f}", orders, f"${delivery:,.2f}"])

    ws1.append(["TOTAL", "$298,000.00", "$286,000.00", "11,400", "$72,000.00"])

    # Tab 2: Labor
    ws2 = wb.create_sheet("Labor Report")
    ws2.append(["Week Of", "Team Member", "Position", "Reg Hours", "OT Hours", "Total Pay"])

    employees = ["Maria Garcia", "James Wilson", "Sam Patel", "Alex Kim",
                 "Jordan Brown", "Casey Davis", "Morgan Lee", "Riley Clark"]

    for week in range(13):
        week_start = start + timedelta(weeks=week)
        for emp in random.sample(employees, random.randint(5, 8)):
            role = "Kitchen" if employees.index(emp) < 4 else "Server"
            reg = round(random.uniform(28, 40), 1)
            ot = round(random.uniform(0, 8), 1) if random.random() < 0.3 else 0
            rate = 18.0 if role == "Kitchen" else 15.0
            pay = round((reg + ot * 1.5) * rate, 2)
            ws2.append([week_start.strftime("%m/%d/%Y"), emp, role, reg, ot, f"${pay:,.2f}"])

    # Tab 3: Refunds & Voids
    ws3 = wb.create_sheet("Refunds Voids Comps")
    ws3.append(["Timestamp", "Server", "Type", "Amount", "Check #", "Manager", "Reason"])

    suspicious = "Sam Patel"
    for i in range(90):
        d = start + timedelta(days=i)
        # Normal refunds
        for _ in range(random.randint(1, 3)):
            emp = random.choice(employees)
            ts = datetime(d.year, d.month, d.day, random.randint(11, 22), random.randint(0, 59))
            amt = round(random.uniform(8, 45), 2)
            rtype = random.choice(["Refund", "Void", "Comp"])
            ws3.append([ts.strftime("%m/%d/%Y %I:%M %p"), emp, rtype, f"${amt:.2f}",
                        f"CHK-{random.randint(1000, 9999)}", "", random.choice(["Wrong item", "Cold", "Customer request", ""])])
        # Suspicious employee extras
        for _ in range(random.randint(2, 4)):
            ts = datetime(d.year, d.month, d.day, random.randint(17, 22), random.randint(0, 59))
            amt = round(random.uniform(15, 55), 2)
            ws3.append([ts.strftime("%m/%d/%Y %I:%M %p"), suspicious, random.choice(["Refund", "Void"]),
                        f"${amt:.2f}", f"CHK-{random.randint(1000, 9999)}", "", ""])

    # Tab 4: Menu Mix
    ws4 = wb.create_sheet("Menu Performance")
    ws4.merge_cells("A1:F1")
    ws4["A1"] = "Menu Mix Report — Q1 2025"
    ws4["A1"].font = Font(bold=True, size=12)
    ws4.append([])
    ws4.append(["Item", "Department", "Qty Sold", "Sales $", "Food Cost %", "Margin $"])

    items = [
        ("Classic Burger", "Entrees", 2840, 39760, "32%", 27037),
        ("Grilled Chicken", "Entrees", 2100, 25200, "28%", 18144),
        ("Caesar Salad", "Entrees", 1560, 14040, "22%", 10951),
        ("French Fries", "Sides", 4200, 16800, "12%", 14784),
        ("Wings Basket", "Apps", 1650, 18150, "33%", 12161),
        ("Craft Beer", "Drinks", 3100, 21700, "15%", 18445),
        ("Soft Drinks", "Drinks", 5200, 15600, "8%", 14352),
        ("Cheesecake", "Desserts", 920, 7360, "20%", 5888),
    ]
    for name, dept, qty, rev, cost_pct, margin in items:
        ws4.append([name, dept, qty, f"${rev:,}", cost_pct, f"${margin:,}"])

    # Tab 5: Random junk (should be ignored/classified as unknown)
    ws5 = wb.create_sheet("Notes")
    ws5.append(["Manager notes — do not share"])
    ws5.append(["Sam might be stealing. Watch the refund pattern."])
    ws5.append(["Need to fix the fryer before Friday."])

    wb.save(OUTPUT)
    print(f"Multi-tab XLSX written to: {OUTPUT}")


if __name__ == "__main__":
    main()
