"""XLSX parser — splits workbooks into sheets, handles merged cells and junk."""

from __future__ import annotations

from pathlib import Path

import openpyxl


def parse_xlsx(file_path: Path) -> dict[str, list[list[str]]]:
    """Read an XLSX file and return {sheet_name: rows} where each row is a list of strings.

    Handles:
    - Multiple sheets
    - Merged cells (unmerges and fills values)
    - None → empty string normalization
    """
    wb = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
    sheets: dict[str, list[list[str]]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Unmerge cells and propagate values
        for merge_range in list(ws.merged_cells.ranges):
            min_row = merge_range.min_row
            min_col = merge_range.min_col
            value = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(str(merge_range))
            for row in range(merge_range.min_row, merge_range.max_row + 1):
                for col in range(merge_range.min_col, merge_range.max_col + 1):
                    ws.cell(row=row, column=col).value = value

        rows: list[list[str]] = []
        for row in ws.iter_rows():
            cells = []
            for cell in row:
                val = cell.value
                if val is None:
                    cells.append("")
                elif isinstance(val, (int, float)):
                    cells.append(str(val))
                else:
                    cells.append(str(val).strip())
            rows.append(cells)

        # Skip completely empty sheets
        if any(any(c for c in row) for row in rows):
            sheets[sheet_name] = rows

    wb.close()
    return sheets
